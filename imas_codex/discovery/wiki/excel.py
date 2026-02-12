"""Excel text extraction with structure-aware formatting.

Provides two extraction modes:
- Preview: Fast extraction of first few sheets/rows for LLM scoring
- Full: Complete extraction with header context for chunking and embedding

Key optimizations over raw cell dumps:
- Header row detection: First non-empty row becomes column headers
- Contextual rows: Data rows formatted as "column: value" pairs
- Numeric sheet skipping: Sheets with < 10% text cells are skipped
- Sheet labeling: Each sheet is clearly delimited with its name
"""

from __future__ import annotations

import logging
import tempfile
from pathlib import Path

logger = logging.getLogger(__name__)

# Minimum fraction of text (non-numeric, non-empty) cells for a sheet
# to be considered worth extracting. Sheets below this threshold are
# purely numeric grids with no semantic value for embeddings.
MIN_TEXT_CELL_FRACTION = 0.10

# Maximum rows to read during preview extraction (scoring)
PREVIEW_MAX_ROWS = 30
PREVIEW_MAX_SHEETS = 3

# Maximum rows to emit per sheet during full extraction before
# switching to a summary. Prevents runaway output on huge data sheets.
FULL_MAX_ROWS_PER_SHEET = 500


def _is_text_value(value: object) -> bool:
    """Check if a cell value is textual (not purely numeric or empty)."""
    if value is None:
        return False
    s = str(value).strip()
    if not s:
        return False
    # Try to parse as a number — if it succeeds, it's not text
    try:
        float(s.replace(",", ""))
        return False
    except (ValueError, OverflowError):
        return True


def _compute_text_fraction(
    rows: list[tuple],
    max_sample: int = 100,
) -> float:
    """Compute fraction of cells that contain text in a sample of rows.

    Samples up to `max_sample` rows to avoid scanning huge sheets.
    """
    total_cells = 0
    text_cells = 0
    for row in rows[:max_sample]:
        for cell in row:
            if cell is not None:
                total_cells += 1
                if _is_text_value(cell):
                    text_cells += 1
    if total_cells == 0:
        return 0.0
    return text_cells / total_cells


def _detect_header_row(rows: list[tuple]) -> tuple[list[str] | None, int]:
    """Detect the header row in a list of rows.

    The header row is the first row where the majority of non-empty cells
    are text (not numbers). Returns (headers, data_start_index).

    If no suitable header is found, returns (None, 0).
    """
    for i, row in enumerate(rows[:5]):  # Check first 5 rows
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if not non_empty:
            continue
        text_count = sum(1 for c in non_empty if _is_text_value(c))
        if len(non_empty) >= 2 and text_count / len(non_empty) >= 0.5:
            headers = [str(c).strip() if c is not None else "" for c in row]
            return headers, i + 1
    return None, 0


def _format_row_with_headers(
    row: tuple,
    headers: list[str] | None,
) -> str:
    """Format a data row, optionally with column headers as context.

    With headers:  "col1: val1 | col2: val2"
    Without:       "val1 | val2"
    """
    if headers:
        parts = []
        for j, cell in enumerate(row):
            if cell is None:
                continue
            val = str(cell).strip()
            if not val:
                continue
            col_name = headers[j] if j < len(headers) and headers[j] else f"col{j + 1}"
            parts.append(f"{col_name}: {val}")
        return " | ".join(parts)
    else:
        values = [str(c).strip() for c in row if c is not None and str(c).strip()]
        return " | ".join(values)


def _is_legacy_xls(content_bytes: bytes) -> bool:
    """Detect legacy .xls (BIFF/Compound Document) format by magic bytes.

    Legacy .xls files start with the OLE2 compound document magic bytes,
    while .xlsx files start with the ZIP magic (PK\x03\x04). openpyxl
    only handles .xlsx — legacy .xls requires xlrd.
    """
    # OLE2 Compound Document magic bytes
    return content_bytes[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _extract_xlrd_sheets(content_bytes: bytes) -> list[tuple[str, list[tuple]]]:
    """Extract sheet names and row tuples from a legacy .xls file using xlrd.

    Returns:
        List of (sheet_name, rows) where rows is a list of tuples.
    """
    import xlrd

    wb = xlrd.open_workbook(file_contents=content_bytes)
    sheets = []
    for sheet in wb.sheets():
        rows = []
        for row_idx in range(sheet.nrows):
            row_values = tuple(
                sheet.cell_value(row_idx, col) for col in range(sheet.ncols)
            )
            rows.append(row_values)
        sheets.append((sheet.name, rows))
    return sheets


def extract_excel_preview(content_bytes: bytes) -> str:
    """Extract a text preview from Excel bytes for LLM scoring.

    Reads up to PREVIEW_MAX_SHEETS sheets, PREVIEW_MAX_ROWS rows each.
    Detects headers and formats rows with column context.
    Skips purely numeric sheets.

    Supports both legacy .xls (via xlrd) and modern .xlsx (via openpyxl).

    Args:
        content_bytes: Raw .xlsx/.xls file bytes

    Returns:
        Text preview string (may be empty if no textual content)
    """
    if _is_legacy_xls(content_bytes):
        return _extract_excel_preview_xlrd(content_bytes)
    return _extract_excel_preview_openpyxl(content_bytes)


def _extract_excel_preview_xlrd(content_bytes: bytes) -> str:
    """Extract preview from legacy .xls using xlrd."""
    sheets = _extract_xlrd_sheets(content_bytes)
    text_parts: list[str] = []
    sheets_processed = 0

    for sheet_name, all_rows in sheets:
        if sheets_processed >= PREVIEW_MAX_SHEETS:
            break
        if not all_rows:
            continue

        text_fraction = _compute_text_fraction(all_rows)
        if text_fraction < MIN_TEXT_CELL_FRACTION:
            continue

        sheets_processed += 1
        headers, data_start = _detect_header_row(all_rows)

        sheet_lines = [f"[Sheet: {sheet_name}]"]
        if headers:
            sheet_lines.append("Columns: " + " | ".join(h for h in headers if h))

        for row in all_rows[data_start : data_start + PREVIEW_MAX_ROWS]:
            line = _format_row_with_headers(row, headers)
            if line:
                sheet_lines.append(line)

        if len(sheet_lines) > 1:
            text_parts.append("\n".join(sheet_lines))

    return "\n\n".join(text_parts)


def _extract_excel_preview_openpyxl(content_bytes: bytes) -> str:
    """Extract preview from modern .xlsx using openpyxl."""
    from openpyxl import load_workbook

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(content_bytes)
        temp_path = Path(f.name)

    try:
        wb = load_workbook(temp_path, data_only=True)
        text_parts: list[str] = []
        sheets_processed = 0

        for sheet_name in wb.sheetnames:
            if sheets_processed >= PREVIEW_MAX_SHEETS:
                break

            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                continue

            # Check if sheet has enough text content
            text_fraction = _compute_text_fraction(all_rows)
            if text_fraction < MIN_TEXT_CELL_FRACTION:
                continue

            sheets_processed += 1
            headers, data_start = _detect_header_row(all_rows)

            sheet_lines = [f"[Sheet: {sheet_name}]"]
            if headers:
                sheet_lines.append("Columns: " + " | ".join(h for h in headers if h))

            for row in all_rows[data_start : data_start + PREVIEW_MAX_ROWS]:
                line = _format_row_with_headers(row, headers)
                if line:
                    sheet_lines.append(line)

            if len(sheet_lines) > 1:
                text_parts.append("\n".join(sheet_lines))

        return "\n\n".join(text_parts)
    finally:
        temp_path.unlink(missing_ok=True)


def extract_excel_full(content_bytes: bytes) -> str:
    """Extract full text from Excel bytes for chunking and embedding.

    Reads all sheets with textual content. Uses header detection to
    produce contextual "column: value" rows. Skips purely numeric sheets.

    Supports both legacy .xls (via xlrd) and modern .xlsx (via openpyxl).

    Args:
        content_bytes: Raw .xlsx/.xls file bytes

    Returns:
        Full text string (may be empty if no textual content)
    """
    if _is_legacy_xls(content_bytes):
        return _extract_excel_full_xlrd(content_bytes)
    return _extract_excel_full_openpyxl(content_bytes)


def _extract_excel_full_xlrd(content_bytes: bytes) -> str:
    """Extract full text from legacy .xls using xlrd."""
    sheets = _extract_xlrd_sheets(content_bytes)
    text_parts: list[str] = []

    for sheet_name, all_rows in sheets:
        if not all_rows:
            continue

        text_fraction = _compute_text_fraction(all_rows)
        if text_fraction < MIN_TEXT_CELL_FRACTION:
            logger.debug(
                "Skipping sheet '%s': only %.0f%% text cells",
                sheet_name,
                text_fraction * 100,
            )
            continue

        headers, data_start = _detect_header_row(all_rows)

        sheet_lines = [f"[Sheet: {sheet_name}]"]
        if headers:
            sheet_lines.append("Columns: " + " | ".join(h for h in headers if h))

        rows_emitted = 0
        for row in all_rows[data_start:]:
            line = _format_row_with_headers(row, headers)
            if line:
                sheet_lines.append(line)
                rows_emitted += 1
                if rows_emitted >= FULL_MAX_ROWS_PER_SHEET:
                    remaining = len(all_rows) - data_start - rows_emitted
                    if remaining > 0:
                        sheet_lines.append(f"[... {remaining} more rows truncated]")
                    break

        if len(sheet_lines) > 1:
            text_parts.append("\n".join(sheet_lines))

    return "\n\n".join(text_parts)


def _extract_excel_full_openpyxl(content_bytes: bytes) -> str:
    """Extract full text from modern .xlsx using openpyxl."""
    from openpyxl import load_workbook

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(content_bytes)
        temp_path = Path(f.name)

    try:
        wb = load_workbook(temp_path, data_only=True)
        text_parts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                continue

            # Check if sheet has enough text content
            text_fraction = _compute_text_fraction(all_rows)
            if text_fraction < MIN_TEXT_CELL_FRACTION:
                logger.debug(
                    "Skipping sheet '%s': only %.0f%% text cells",
                    sheet_name,
                    text_fraction * 100,
                )
                continue

            headers, data_start = _detect_header_row(all_rows)

            sheet_lines = [f"[Sheet: {sheet_name}]"]
            if headers:
                sheet_lines.append("Columns: " + " | ".join(h for h in headers if h))

            rows_emitted = 0
            for row in all_rows[data_start:]:
                line = _format_row_with_headers(row, headers)
                if line:
                    sheet_lines.append(line)
                    rows_emitted += 1
                    if rows_emitted >= FULL_MAX_ROWS_PER_SHEET:
                        remaining = len(all_rows) - data_start - rows_emitted
                        if remaining > 0:
                            sheet_lines.append(f"[... {remaining} more rows truncated]")
                        break

            if len(sheet_lines) > 1:
                text_parts.append("\n".join(sheet_lines))

        return "\n\n".join(text_parts)
    finally:
        temp_path.unlink(missing_ok=True)
