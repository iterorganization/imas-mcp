"""Tests for Excel text extraction with structure-aware formatting."""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook

from imas_codex.discovery.wiki.excel import (
    FULL_MAX_ROWS_PER_SHEET,
    MIN_TEXT_CELL_FRACTION,
    PREVIEW_MAX_ROWS,
    PREVIEW_MAX_SHEETS,
    _compute_text_fraction,
    _detect_header_row,
    _format_row_with_headers,
    _is_legacy_xls,
    _is_text_value,
    extract_excel_full,
    extract_excel_preview,
)

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_xlsx(*sheets: tuple[str, list[list]]) -> bytes:
    """Create an in-memory xlsx from (sheet_name, rows) pairs."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    for name, rows in sheets:
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# _is_text_value
# ---------------------------------------------------------------------------


class TestIsTextValue:
    def test_none_is_not_text(self):
        assert _is_text_value(None) is False

    def test_empty_string_is_not_text(self):
        assert _is_text_value("") is False
        assert _is_text_value("   ") is False

    def test_integer_string_is_not_text(self):
        assert _is_text_value("42") is False
        assert _is_text_value("3.14") is False

    def test_comma_number_is_not_text(self):
        assert _is_text_value("1,234") is False

    def test_word_is_text(self):
        assert _is_text_value("electron") is True
        assert _is_text_value("Te [keV]") is True

    def test_mixed_alphanumeric_is_text(self):
        assert _is_text_value("shot_12345") is True


# ---------------------------------------------------------------------------
# _compute_text_fraction
# ---------------------------------------------------------------------------


class TestComputeTextFraction:
    def test_all_text(self):
        rows = [("alpha", "beta"), ("gamma", "delta")]
        assert _compute_text_fraction(rows) == 1.0

    def test_all_numeric(self):
        rows = [(1, 2, 3), (4, 5, 6)]
        assert _compute_text_fraction(rows) == 0.0

    def test_mixed(self):
        rows = [("header", 1), ("label", 2)]
        frac = _compute_text_fraction(rows)
        assert 0.4 < frac < 0.6  # 2 text out of 4

    def test_empty_rows(self):
        rows = [(None, None), (None, None)]
        assert _compute_text_fraction(rows) == 0.0

    def test_empty_list(self):
        assert _compute_text_fraction([]) == 0.0


# ---------------------------------------------------------------------------
# _detect_header_row
# ---------------------------------------------------------------------------


class TestDetectHeaderRow:
    def test_typical_header(self):
        rows = [
            ("Name", "Value", "Unit"),
            ("Te", 1.5, "keV"),
            ("ne", 2e19, "m^-3"),
        ]
        headers, start = _detect_header_row(rows)
        assert headers == ["Name", "Value", "Unit"]
        assert start == 1

    def test_no_header_all_numeric(self):
        rows = [(1, 2, 3), (4, 5, 6)]
        headers, start = _detect_header_row(rows)
        assert headers is None
        assert start == 0

    def test_blank_first_row_then_header(self):
        rows = [
            (None, None),
            ("Parameter", "Description"),
            ("ip", "Plasma current"),
        ]
        headers, start = _detect_header_row(rows)
        assert headers == ["Parameter", "Description"]
        assert start == 2

    def test_single_text_column_not_enough(self):
        """Need at least 2 non-empty cells for a header row."""
        rows = [("Title", None, None), (1, 2, 3)]
        headers, start = _detect_header_row(rows)
        # Only 1 non-empty cell — not a header
        assert headers is None
        assert start == 0


# ---------------------------------------------------------------------------
# _format_row_with_headers
# ---------------------------------------------------------------------------


class TestFormatRowWithHeaders:
    def test_with_headers(self):
        result = _format_row_with_headers(
            ("Te", 1.5, "keV"),
            ["Name", "Value", "Unit"],
        )
        assert "Name: Te" in result
        assert "Value: 1.5" in result
        assert "Unit: keV" in result

    def test_without_headers(self):
        result = _format_row_with_headers(("Te", 1.5, "keV"), None)
        assert result == "Te | 1.5 | keV"

    def test_skips_none(self):
        result = _format_row_with_headers(("Te", None, "keV"), None)
        assert result == "Te | keV"

    def test_header_shorter_than_row(self):
        result = _format_row_with_headers(
            ("a", "b", "c"),
            ["Col1", "Col2"],  # Only 2 headers for 3 values
        )
        assert "Col1: a" in result
        assert "Col2: b" in result
        assert "col3: c" in result  # Falls back to colN


# ---------------------------------------------------------------------------
# extract_excel_preview
# ---------------------------------------------------------------------------


class TestExtractExcelPreview:
    def test_basic_table(self):
        data = _make_xlsx(
            (
                "Params",
                [
                    ["Name", "Value", "Unit"],
                    ["Te", 1.5, "keV"],
                    ["ne", 2e19, "m^-3"],
                ],
            )
        )
        text = extract_excel_preview(data)
        assert "[Sheet: Params]" in text
        assert "Columns:" in text
        assert "Name: Te" in text

    def test_skips_numeric_sheet(self):
        """Sheets with < MIN_TEXT_CELL_FRACTION text cells are skipped."""
        data = _make_xlsx(
            ("TextSheet", [["Name", "Desc"], ["ip", "Current"]]),
            ("NumberSheet", [[i, i * 2, i * 3] for i in range(20)]),
        )
        text = extract_excel_preview(data)
        assert "TextSheet" in text
        assert "NumberSheet" not in text

    def test_max_sheets_limit(self):
        """Only PREVIEW_MAX_SHEETS sheets are processed."""
        sheets = []
        for i in range(PREVIEW_MAX_SHEETS + 2):
            sheets.append((f"Sheet{i}", [["Header", "Data"], ["val", "info"]]))
        data = _make_xlsx(*sheets)
        text = extract_excel_preview(data)
        # Count sheet headers — should be at most PREVIEW_MAX_SHEETS
        assert text.count("[Sheet:") <= PREVIEW_MAX_SHEETS

    def test_max_rows_limit(self):
        """Only PREVIEW_MAX_ROWS data rows are included per sheet."""
        rows = [["Name", "Value"]] + [[f"param_{i}", "data"] for i in range(50)]
        data = _make_xlsx(("Big", rows))
        text = extract_excel_preview(data)
        lines = [line for line in text.split("\n") if line.startswith("Name:")]
        assert len(lines) <= PREVIEW_MAX_ROWS

    def test_empty_workbook(self):
        data = _make_xlsx(("Empty", []))
        text = extract_excel_preview(data)
        assert text == ""


# ---------------------------------------------------------------------------
# extract_excel_full
# ---------------------------------------------------------------------------


class TestExtractExcelFull:
    def test_basic_table(self):
        data = _make_xlsx(
            (
                "Results",
                [
                    ["Parameter", "Value", "Unit"],
                    ["Te", 1.5, "keV"],
                    ["ne", 2e19, "m^-3"],
                ],
            )
        )
        text = extract_excel_full(data)
        assert "[Sheet: Results]" in text
        assert "Columns:" in text
        assert "Parameter: Te" in text

    def test_all_sheets_processed(self):
        """Full extraction reads all sheets (not limited like preview)."""
        sheets = [(f"S{i}", [["A", "B"], ["x", "y"]]) for i in range(5)]
        data = _make_xlsx(*sheets)
        text = extract_excel_full(data)
        assert text.count("[Sheet:") == 5

    def test_truncation_of_large_sheet(self):
        """Sheets with more than FULL_MAX_ROWS_PER_SHEET rows are truncated."""
        rows = [["Name", "Value"]] + [
            [f"item_{i}", f"desc_{i}"] for i in range(FULL_MAX_ROWS_PER_SHEET + 50)
        ]
        data = _make_xlsx(("Huge", rows))
        text = extract_excel_full(data)
        assert "more rows truncated" in text

    def test_skips_numeric_sheet(self):
        data = _make_xlsx(
            ("Text", [["Name", "Desc"], ["ip", "Current"]]),
            ("Numbers", [[i * 0.1, i * 0.2] for i in range(20)]),
        )
        text = extract_excel_full(data)
        assert "Text" in text
        assert "Numbers" not in text

    def test_empty_workbook(self):
        data = _make_xlsx(("Empty", []))
        text = extract_excel_full(data)
        assert text == ""

    def test_multiple_sheets_with_mixed_content(self):
        """Full extraction handles mixed text/numeric sheets correctly."""
        data = _make_xlsx(
            (
                "Config",
                [["Setting", "Value"], ["resolution", "high"], ["mode", "auto"]],
            ),
            ("Calibration", [[0.1, 0.2, 0.3], [0.4, 0.5, 0.6]]),
            ("Notes", [["Topic", "Comment"], ["IMAS", "v3.42"], ["MDSplus", "trunk"]]),
        )
        text = extract_excel_full(data)
        assert "Config" in text
        assert "Notes" in text
        # Calibration is all-numeric, should be skipped
        assert "Calibration" not in text


# ---------------------------------------------------------------------------
# Legacy .xls format (xlrd)
# ---------------------------------------------------------------------------


def _make_xls(*sheets: tuple[str, list[list]]) -> bytes:
    """Create an in-memory legacy .xls from (sheet_name, rows) pairs using xlrd's companion xlwt."""
    import xlwt

    wb = xlwt.Workbook()
    for name, rows in sheets:
        ws = wb.add_sheet(name)
        for r_idx, row in enumerate(rows):
            for c_idx, val in enumerate(row):
                ws.write(r_idx, c_idx, val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestIsLegacyXls:
    def test_xlsx_is_not_legacy(self):
        data = _make_xlsx(("Sheet1", [["a", "b"]]))
        assert _is_legacy_xls(data) is False

    def test_xls_is_legacy(self):
        data = _make_xls(("Sheet1", [["a", "b"]]))
        assert _is_legacy_xls(data) is True

    def test_empty_bytes(self):
        assert _is_legacy_xls(b"") is False

    def test_random_bytes(self):
        assert _is_legacy_xls(b"PK\x03\x04random") is False


class TestLegacyXlsPreview:
    def test_basic_xls_table(self):
        data = _make_xls(
            (
                "Params",
                [
                    ["Name", "Value", "Unit"],
                    ["Te", 1.5, "keV"],
                    ["ne", 2e19, "m^-3"],
                ],
            )
        )
        text = extract_excel_preview(data)
        assert "[Sheet: Params]" in text
        assert "Name: Te" in text

    def test_xls_skips_numeric_sheet(self):
        data = _make_xls(
            ("TextSheet", [["Name", "Desc"], ["ip", "Current"]]),
            ("NumberSheet", [[i, i * 2, i * 3] for i in range(20)]),
        )
        text = extract_excel_preview(data)
        assert "TextSheet" in text
        assert "NumberSheet" not in text


class TestLegacyXlsFull:
    def test_basic_xls_full(self):
        data = _make_xls(
            (
                "Results",
                [
                    ["Parameter", "Value", "Unit"],
                    ["Te", 1.5, "keV"],
                    ["ne", 2e19, "m^-3"],
                ],
            )
        )
        text = extract_excel_full(data)
        assert "[Sheet: Results]" in text
        assert "Parameter: Te" in text

    def test_xls_multiple_sheets(self):
        data = _make_xls(
            ("Config", [["Setting", "Value"], ["resolution", "high"]]),
            ("Notes", [["Topic", "Comment"], ["IMAS", "v3.42"]]),
        )
        text = extract_excel_full(data)
        assert "Config" in text
        assert "Notes" in text
